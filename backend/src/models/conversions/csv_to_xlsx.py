import os
import logging
import pandas as pd
from pathlib import Path

# Importar librerías para Excel de alta calidad
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl no disponible para CSV→XLSX de alta calidad")

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False
    logging.warning("xlsxwriter no disponible para CSV→XLSX")

CONVERSION = ('csv', 'xlsx')

def convert(input_path, output_path):
    """Convierte CSV a XLSX usando la mejor librería disponible"""
    
    # Método 1: pandas + openpyxl (RECOMENDADO - mejor calidad y formato)
    if OPENPYXL_AVAILABLE:
        try:
            success, message = convert_with_openpyxl(input_path, output_path)
            if success:
                return True, f"Conversión CSV→XLSX exitosa con openpyxl - {message}"
        except Exception as e:
            logging.warning(f"openpyxl falló: {e}")
    
    # Método 2: pandas + xlsxwriter (buena calidad)
    if XLSXWRITER_AVAILABLE:
        try:
            success, message = convert_with_xlsxwriter(input_path, output_path)
            if success:
                return True, f"Conversión CSV→XLSX exitosa con xlsxwriter - {message}"
        except Exception as e:
            logging.warning(f"xlsxwriter falló: {e}")
    
    # Método 3: pandas básico (fallback)
    return convert_with_pandas_basic(input_path, output_path)

def convert_with_openpyxl(input_path, output_path):
    """Conversión usando pandas + openpyxl (máxima calidad)"""
    try:
        # Detectar encoding del CSV
        encoding = detect_csv_encoding(input_path)
        
        # Leer CSV con pandas
        df = pd.read_csv(input_path, encoding=encoding)
        
        if df.empty:
            return False, "CSV vacío o sin datos válidos"
        
        # Crear workbook con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos CSV"
        
        # Agregar datos del DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Aplicar formato profesional
        apply_professional_formatting(ws, df)
        
        # Guardar archivo Excel
        wb.save(output_path)
        
        # Verificar resultado
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            rows, cols = df.shape
            return True, f"Excel generado: {rows} filas, {cols} columnas con formato profesional"
        else:
            return False, "Error: XLSX no se generó correctamente"
        
    except Exception as e:
        return False, f"Error con openpyxl: {str(e)}"

def convert_with_xlsxwriter(input_path, output_path):
    """Conversión usando pandas + xlsxwriter"""
    try:
        # Detectar encoding del CSV
        encoding = detect_csv_encoding(input_path)
        
        # Leer CSV con pandas
        df = pd.read_csv(input_path, encoding=encoding)
        
        if df.empty:
            return False, "CSV vacío o sin datos válidos"
        
        # Crear Excel con xlsxwriter
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Datos CSV', index=False)
            
            # Obtener workbook y worksheet para formato
            workbook = writer.book
            worksheet = writer.sheets['Datos CSV']
            
            # Aplicar formato básico
            apply_xlsxwriter_formatting(workbook, worksheet, df)
        
        # Verificar resultado
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            rows, cols = df.shape
            return True, f"Excel generado: {rows} filas, {cols} columnas con xlsxwriter"
        else:
            return False, "Error: XLSX no se generó correctamente"
        
    except Exception as e:
        return False, f"Error con xlsxwriter: {str(e)}"

def convert_with_pandas_basic(input_path, output_path):
    """Conversión básica usando solo pandas"""
    try:
        # Detectar encoding del CSV
        encoding = detect_csv_encoding(input_path)
        
        # Leer CSV con pandas
        df = pd.read_csv(input_path, encoding=encoding)
        
        if df.empty:
            return False, "CSV vacío o sin datos válidos"
        
        # Guardar como Excel básico
        df.to_excel(output_path, sheet_name='Datos CSV', index=False, engine='openpyxl')
        
        # Verificar resultado
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            rows, cols = df.shape
            return True, f"Excel básico generado: {rows} filas, {cols} columnas"
        else:
            return False, "Error: XLSX no se generó correctamente"
        
    except Exception as e:
        return False, f"Error con pandas básico: {str(e)}"

def detect_csv_encoding(file_path):
    """Detectar encoding del archivo CSV"""
    try:
        import chardet
        
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)  # Leer primeros 10KB
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            
            # Mapear encodings comunes
            encoding_map = {
                'ascii': 'utf-8',
                'ISO-8859-1': 'latin-1',
                'Windows-1252': 'cp1252'
            }
            
            return encoding_map.get(encoding, encoding or 'utf-8')
            
    except ImportError:
        # Si chardet no está disponible, probar encodings comunes
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1000)  # Probar leer
                return encoding
            except UnicodeDecodeError:
                continue
        
        return 'utf-8'  # Fallback por defecto

def apply_professional_formatting(worksheet, dataframe):
    """Aplicar formato profesional al worksheet con openpyxl"""
    try:
        # Definir estilos
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formato a headers (primera fila)
        for col in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Aplicar formato a datos
        for row in range(2, len(dataframe) + 2):
            for col in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Ajustar ancho de columnas automáticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila (headers)
        worksheet.freeze_panes = 'A2'
        
    except Exception as e:
        logging.warning(f"Error aplicando formato profesional: {e}")

def apply_xlsxwriter_formatting(workbook, worksheet, dataframe):
    """Aplicar formato básico con xlsxwriter"""
    try:
        # Definir formatos
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Aplicar formato a headers
        for col, column_name in enumerate(dataframe.columns):
            worksheet.write(0, col, column_name, header_format)
        
        # Ajustar ancho de columnas
        for col, column_name in enumerate(dataframe.columns):
            max_length = max(
                len(str(column_name)),
                dataframe[column_name].astype(str).str.len().max()
            )
            worksheet.set_column(col, col, min(max_length + 2, 50))
        
        # Congelar primera fila
        worksheet.freeze_panes(1, 0)
        
    except Exception as e:
        logging.warning(f"Error aplicando formato xlsxwriter: {e}")
