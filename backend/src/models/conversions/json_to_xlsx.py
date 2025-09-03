import os
import logging
import pandas as pd
import json
from pathlib import Path

# Importar librerías para Excel de alta calidad
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl no disponible para JSON→XLSX de alta calidad")

CONVERSION = ('json', 'xlsx')

def convert(input_path, output_path):
    """Convierte JSON a XLSX usando pandas con formato profesional"""
    
    try:
        # Leer y validar JSON
        with open(input_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        if not json_data:
            return False, "JSON vacío o sin datos válidos"
        
        # Determinar estructura del JSON y convertir a DataFrame(s)
        dataframes = json_to_dataframes(json_data)
        
        if not dataframes:
            return False, "No se pudieron extraer datos tabulares del JSON"
        
        # Crear Excel con formato profesional
        if OPENPYXL_AVAILABLE:
            success, message = create_excel_with_openpyxl(dataframes, output_path)
        else:
            success, message = create_excel_basic(dataframes, output_path)
        
        if success:
            return True, f"Excel generado desde JSON - {message}"
        else:
            return False, f"Error creando Excel: {message}"
        
    except json.JSONDecodeError as e:
        return False, f"JSON inválido: {str(e)}"
    except Exception as e:
        return False, f"Error en conversión JSON→XLSX: {str(e)}"

def json_to_dataframes(json_data):
    """Convertir JSON a uno o más DataFrames"""
    try:
        dataframes = {}
        
        # Verificar si tiene estructura de metadatos
        if isinstance(json_data, dict) and 'data' in json_data and 'metadata' in json_data:
            json_data = json_data['data']
        
        # Caso 1: Lista de objetos [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
        if isinstance(json_data, list):
            df = handle_json_list(json_data)
            if not df.empty:
                dataframes['Datos'] = df
        
        # Caso 2: Objeto con múltiples estructuras (posibles hojas)
        elif isinstance(json_data, dict):
            # Verificar si es formato columnar {"col1": [1, 2], "col2": [3, 4]}
            if all(isinstance(value, list) for value in json_data.values()):
                lengths = [len(value) for value in json_data.values()]
                if len(set(lengths)) == 1:  # Todas las longitudes son iguales
                    df = pd.DataFrame(json_data)
                    if not df.empty:
                        dataframes['Datos'] = df
            
            # Verificar si cada clave puede ser una hoja separada
            else:
                for key, value in json_data.items():
                    df = None
                    
                    if isinstance(value, list):
                        df = handle_json_list(value)
                    elif isinstance(value, dict):
                        # Si es un objeto con arrays, intentar convertir
                        if all(isinstance(v, list) for v in value.values()):
                            lengths = [len(v) for v in value.values()]
                            if len(set(lengths)) == 1:
                                df = pd.DataFrame(value)
                    
                    if df is not None and not df.empty:
                        clean_key = clean_sheet_name(key)
                        dataframes[clean_key] = df
                
                # Si no se encontraron hojas, intentar convertir el objeto completo
                if not dataframes:
                    df = pd.json_normalize([json_data])
                    if not df.empty:
                        dataframes['Datos'] = df
        
        return dataframes
        
    except Exception as e:
        logging.error(f"Error convirtiendo JSON a DataFrames: {e}")
        return {}

def handle_json_list(json_list):
    """Manejar JSON que es una lista"""
    try:
        if not json_list:
            return pd.DataFrame()
        
        # Si todos los elementos son objetos/diccionarios
        if all(isinstance(item, dict) for item in json_list):
            # Usar pandas.json_normalize para aplanar objetos anidados
            df = pd.json_normalize(json_list)
            return df
        
        # Si todos los elementos son listas
        elif all(isinstance(item, list) for item in json_list):
            # Crear DataFrame con columnas numeradas
            max_cols = max(len(item) for item in json_list) if json_list else 0
            columns = [f'columna_{i+1}' for i in range(max_cols)]
            
            # Rellenar listas cortas con None
            padded_data = []
            for item in json_list:
                padded_item = item + [None] * (max_cols - len(item))
                padded_data.append(padded_item)
            
            return pd.DataFrame(padded_data, columns=columns)
        
        # Si son valores simples
        else:
            return pd.DataFrame({'valor': json_list})
        
    except Exception as e:
        logging.error(f"Error manejando lista JSON: {e}")
        return pd.DataFrame()

def create_excel_with_openpyxl(dataframes, output_path):
    """Crear Excel con formato profesional usando openpyxl"""
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        for sheet_name, df in dataframes.items():
            # Crear nueva hoja
            ws = wb.create_sheet(title=sheet_name)
            
            # Limpiar datos para Excel
            df_cleaned = clean_dataframe_for_excel(df)
            
            # Agregar datos del DataFrame
            for r in dataframe_to_rows(df_cleaned, index=False, header=True):
                ws.append(r)
            
            # Aplicar formato profesional
            apply_professional_formatting(ws, df_cleaned)
        
        # Guardar archivo Excel
        wb.save(output_path)
        
        # Verificar resultado
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            total_rows = sum(len(df) for df in dataframes.values())
            sheet_count = len(dataframes)
            return True, f"{total_rows} filas en {sheet_count} hoja(s) con formato profesional"
        else:
            return False, "XLSX no se generó correctamente"
        
    except Exception as e:
        return False, f"Error con openpyxl: {str(e)}"

def create_excel_basic(dataframes, output_path):
    """Crear Excel básico usando pandas"""
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in dataframes.items():
                # Limpiar datos para Excel
                df_cleaned = clean_dataframe_for_excel(df)
                
                # Escribir a Excel
                df_cleaned.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Verificar resultado
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            total_rows = sum(len(df) for df in dataframes.values())
            sheet_count = len(dataframes)
            return True, f"{total_rows} filas en {sheet_count} hoja(s) básico"
        else:
            return False, "XLSX no se generó correctamente"
        
    except Exception as e:
        return False, f"Error con pandas básico: {str(e)}"

def clean_dataframe_for_excel(df):
    """Limpiar DataFrame para exportación Excel"""
    try:
        # Crear copia para no modificar original
        df_clean = df.copy()
        
        # Eliminar filas completamente vacías
        df_clean = df_clean.dropna(how='all')
        
        # Eliminar columnas completamente vacías
        df_clean = df_clean.dropna(axis=1, how='all')
        
        # Limpiar nombres de columnas
        df_clean.columns = [clean_column_name(str(col)) for col in df_clean.columns]
        
        # Limpiar datos de celdas
        for column in df_clean.columns:
            df_clean[column] = df_clean[column].apply(clean_cell_value_for_excel)
        
        return df_clean
        
    except Exception as e:
        logging.warning(f"Error limpiando DataFrame: {e}")
        return df

def clean_sheet_name(name):
    """Limpiar nombre de hoja para Excel"""
    try:
        # Convertir a string
        clean_name = str(name)
        
        # Caracteres no permitidos en nombres de hoja de Excel
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        # Limitar longitud (Excel permite máximo 31 caracteres)
        if len(clean_name) > 31:
            clean_name = clean_name[:28] + '...'
        
        # No puede estar vacío
        if not clean_name.strip():
            clean_name = 'Hoja'
        
        return clean_name.strip()
        
    except Exception:
        return 'Hoja'

def clean_column_name(column_name):
    """Limpiar nombre de columna para Excel"""
    try:
        # Convertir a string si no lo es
        name = str(column_name)
        
        # Reemplazar valores problemáticos
        if name.lower() in ['unnamed', 'nan', 'none', '']:
            return 'Columna_Sin_Nombre'
        
        # Limpiar caracteres problemáticos
        name = name.replace('\n', ' ').replace('\r', ' ')
        
        # Reemplazar puntos de json_normalize
        name = name.replace('.', '_')
        
        # Limitar longitud
        if len(name) > 50:
            name = name[:47] + '...'
        
        return name.strip()
        
    except Exception:
        return 'Columna_Error'

def clean_cell_value_for_excel(value):
    """Limpiar valor de celda para Excel"""
    try:
        # Manejar valores nulos
        if pd.isna(value) or value is None:
            return ''
        
        # Si es un diccionario o lista, convertir a JSON string
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, separators=(',', ':'))
        
        # Convertir a string
        str_value = str(value)
        
        # Limpiar caracteres problemáticos
        str_value = str_value.replace('\n', ' ').replace('\r', ' ')
        
        # Excel tiene límite de 32,767 caracteres por celda
        if len(str_value) > 32767:
            str_value = str_value[:32764] + '...'
        
        return str_value.strip()
        
    except Exception:
        return str(value) if value is not None else ''

def apply_professional_formatting(worksheet, dataframe):
    """Aplicar formato profesional al worksheet con openpyxl"""
    try:
        # Definir estilos
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formato a headers (primera fila)
        for col in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Aplicar formato a datos
        for row in range(2, len(dataframe) + 2):
            for col in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Ajustar ancho de columnas automáticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila (headers)
        worksheet.freeze_panes = 'A2'
        
    except Exception as e:
        logging.warning(f"Error aplicando formato profesional: {e}")
